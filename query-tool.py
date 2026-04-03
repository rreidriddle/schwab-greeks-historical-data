"""
Greeks Database Query Tool
==========================
Query your greeks_history.db and export results to a formatted Excel file.

QUICK START — just edit the CONFIG section below, then run:
    python query_tool.py

No other files needed. Works with your existing greeks_history.db.
"""

import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import os
import sys

# ══════════════════════════════════════════════════════════════════════════════
#  CONFIG  —  Edit everything in this section
# ══════════════════════════════════════════════════════════════════════════════

DB_PATH = r"D:/GreeksData/greeks_history.db"   # path to your SQLite database

# ── Tickers ───────────────────────────────────────────────────────────────────
# Any combination of: "SPY", "QQQ", "DIA", "XSP", "IWM"
# Use [] for ALL symbols
SYMBOLS = ["SPY", "QQQ"]

# ── DTE filter ────────────────────────────────────────────────────────────────
# Set these to isolate a specific expiry.
# Today (April 3) → 8 May 2026 is ~35 DTE, so 30-40 captures it cleanly.
# Leave as None to skip DTE filtering entirely.
DTE_LOW  = 30
DTE_HIGH = 40

# ── DTE bucket filter ─────────────────────────────────────────────────────────
# Filter by named bucket instead of (or in addition to) raw DTE range.
# Options: "0DTE" | "1-7DTE" | "8-30DTE" | "31-90DTE" | "91-180DTE" | "180+DTE"
# Set to None to skip bucket filtering.
DTE_BUCKET = "31-90DTE"

# ── Date/time range ───────────────────────────────────────────────────────────
# Filters by when your collector captured the snapshot (not the expiry date).
# Format: "YYYY-MM-DD" for a full day, or "YYYY-MM-DD HH:MM:SS" for exact time.
# Set to None to pull all available timestamps.
DATE_FROM = None     # e.g. "2026-04-01"
DATE_TO   = None     # e.g. "2026-04-03"

# ── Snapshot selection ────────────────────────────────────────────────────────
# If multiple pull timestamps exist in your date range, which to use?
# "latest"  → most recent pull only
# "earliest"→ first pull only
# "all"     → every pull in range (can be large)
SNAPSHOT = "latest"

# ── Output ────────────────────────────────────────────────────────────────────
# File saved in same folder as this script.
# Leave as None to auto-generate a name from your filters.
OUTPUT_FILENAME = None    # e.g. "my_query.xlsx"

# ── Tables to export ──────────────────────────────────────────────────────────
# "strike"  → per-strike GEX / Vanna / Charm detail
# "summary" → top-level net exposures, walls, flip levels
# "both"    → separate sheets for each
EXPORT_TABLE = "both"

# ══════════════════════════════════════════════════════════════════════════════
#  END CONFIG  —  nothing below needs changing for normal use
# ══════════════════════════════════════════════════════════════════════════════

DARK      = "1E1E2E"
HDR_BG    = "2D2D44"
ALT_BG    = "252538"
WHITE     = "FFFFFF"
GRAY      = "A0A0C0"
GREEN     = "10B981"
BLUE      = "60A5FA"
AMBER     = "F59E0B"
RED       = "EF4444"
PURPLE    = "A78BFA"

def validate_config():
    errors = []
    if not os.path.exists(DB_PATH):
        errors.append(f"Database not found: {DB_PATH}")
    if EXPORT_TABLE not in ("strike", "summary", "both"):
        errors.append(f"EXPORT_TABLE must be 'strike', 'summary', or 'both'")
    if SNAPSHOT not in ("latest", "earliest", "all"):
        errors.append(f"SNAPSHOT must be 'latest', 'earliest', or 'all'")
    valid_buckets = {"0DTE","1-7DTE","8-30DTE","31-90DTE","91-180DTE","180+DTE",None}
    if DTE_BUCKET not in valid_buckets:
        errors.append(f"DTE_BUCKET '{DTE_BUCKET}' not recognized. Options: {valid_buckets - {None}}")
    if errors:
        print("\n  CONFIG ERRORS:")
        for e in errors:
            print(f"    ✗  {e}")
        sys.exit(1)

def build_filename():
    if OUTPUT_FILENAME:
        return OUTPUT_FILENAME
    syms   = "_".join(SYMBOLS) if SYMBOLS else "ALL"
    bucket = DTE_BUCKET.replace("+","plus") if DTE_BUCKET else f"DTE{DTE_LOW}-{DTE_HIGH}"
    snap   = SNAPSHOT
    date   = datetime.now().strftime("%Y%m%d_%H%M")
    return f"greeks_{syms}_{bucket}_{snap}_{date}.xlsx"

def get_timestamps(conn, symbols):
    sym_filter = ""
    if symbols:
        placeholders = ",".join("?" * len(symbols))
        sym_filter = f"AND symbol IN ({placeholders})"
    params = list(symbols) if symbols else []

    date_clauses = []
    if DATE_FROM:
        date_clauses.append("timestamp >= ?")
        params.append(DATE_FROM)
    if DATE_TO:
        date_clauses.append("timestamp <= ?")
        params.append(DATE_TO + " 23:59:59")
    date_filter = ("AND " + " AND ".join(date_clauses)) if date_clauses else ""

    if SNAPSHOT == "latest":
        order, limit = "DESC", 1
    elif SNAPSHOT == "earliest":
        order, limit = "ASC", 1
    else:
        order, limit = "ASC", -1

    limit_clause = f"LIMIT {limit}" if limit > 0 else ""

    rows = conn.execute(f"""
        SELECT DISTINCT timestamp FROM summary
        WHERE 1=1 {sym_filter} {date_filter}
        ORDER BY timestamp {order}
        {limit_clause}
    """, params).fetchall()

    return [r[0] for r in rows]

def query_strike_data(conn, timestamps, symbols):
    ts_ph   = ",".join("?" * len(timestamps))
    params  = list(timestamps)
    clauses = [f"timestamp IN ({ts_ph})"]

    if symbols:
        sym_ph = ",".join("?" * len(symbols))
        clauses.append(f"symbol IN ({sym_ph})")
        params += list(symbols)
    if DTE_BUCKET:
        clauses.append("dte_bucket = ?")
        params.append(DTE_BUCKET)
    if DTE_LOW is not None:
        clauses.append("dte >= ?")
        params.append(DTE_LOW)
    if DTE_HIGH is not None:
        clauses.append("dte <= ?")
        params.append(DTE_HIGH)

    where = " AND ".join(clauses)
    return pd.read_sql(f"""
        SELECT
            timestamp, symbol, spot, strike, dte, dte_bucket,
            GEX_call, GEX_put, GEX_net,
            VannEX_call, VannEX_put, VannEX_net,
            CharmEX_call, CharmEX_put, CharmEX_net,
            total_oi, total_volume, iv_call, iv_put
        FROM strike_data
        WHERE {where}
        ORDER BY symbol, strike, timestamp
    """, conn, params=params)

def query_summary_data(conn, timestamps, symbols):
    ts_ph  = ",".join("?" * len(timestamps))
    params = list(timestamps)
    clauses = [f"timestamp IN ({ts_ph})"]

    if symbols:
        sym_ph = ",".join("?" * len(symbols))
        clauses.append(f"symbol IN ({sym_ph})")
        params += list(symbols)

    where = " AND ".join(clauses)
    return pd.read_sql(f"""
        SELECT
            timestamp, symbol, spot,
            net_GEX, net_VannEX, net_CharmEX,
            gamma_flip, call_wall, put_wall, max_pain,
            total_oi, total_volume, iv_atm,
            gex_0dte, gex_1_7dte, gex_8_30dte,
            gex_31_90dte, gex_91_180dte, gex_180plus_dte
        FROM summary
        WHERE {where}
        ORDER BY symbol, timestamp
    """, conn, params=params)

# ── Excel styling helpers ─────────────────────────────────────────────────────

def hdr_cell(ws, ref, value, bg=HDR_BG, fc=WHITE, bold=True, size=10, align="center"):
    c = ws[ref]
    c.value = value
    c.font = Font(name="Calibri", bold=bold, color=fc, size=size)
    c.fill = PatternFill("solid", start_color=bg, end_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")

def data_cell(ws, ref, value, fc=WHITE, bg=DARK, align="right", nf=None, bold=False, size=10):
    c = ws[ref]
    c.value = value
    c.font = Font(name="Calibri", color=fc, size=size, bold=bold)
    c.fill = PatternFill("solid", start_color=bg, end_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if nf:
        c.number_format = nf

def thin_border(color="3D3D5C"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def color_for_value(val):
    """Green for positive, red for negative, gray for zero."""
    if val is None or val == 0:
        return GRAY
    return GREEN if val > 0 else RED

def write_title_sheet(wb, symbols, timestamps, df_strike, df_summary):
    ws = wb.create_sheet("Query Info")
    ws.sheet_view.showGridLines = False

    for row in range(1, 30):
        for col in range(1, 10):
            ws[f"{get_column_letter(col)}{row}"].fill = PatternFill("solid", start_color=DARK, end_color=DARK)

    ws.column_dimensions["A"].hidden = True
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 35
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 36

    ws.merge_cells("B2:C2")
    hdr_cell(ws, "B2", "  GREEKS DATABASE QUERY RESULTS", bg="7C3AED", size=15, align="left")

    info = [
        ("Symbols",        ", ".join(symbols) if symbols else "ALL"),
        ("DTE Bucket",     DTE_BUCKET or "—"),
        ("DTE Range",      f"{DTE_LOW} – {DTE_HIGH}" if DTE_LOW is not None else "—"),
        ("Snapshot",       SNAPSHOT),
        ("Date From",      DATE_FROM or "—"),
        ("Date To",        DATE_TO or "—"),
        ("Timestamps",     str(len(timestamps))),
        ("Strike rows",    f"{len(df_strike):,}" if df_strike is not None else "—"),
        ("Summary rows",   f"{len(df_summary):,}" if df_summary is not None else "—"),
        ("Generated",      datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("IV Call/Put",    "NULL in DB — see collector.py write_strike_data() fix needed"),
    ]

    for i, (lbl, val) in enumerate(info):
        r = i + 4
        ws.row_dimensions[r].height = 20
        hdr_cell(ws, f"B{r}", lbl, bg=HDR_BG, align="left", bold=True)
        data_cell(ws, f"C{r}", val, align="left",
                  bg=ALT_BG if i % 2 == 0 else DARK)

def write_strike_sheet(wb, df):
    ws = wb.create_sheet("Strike Data")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].hidden = True

    for row in range(1, len(df) + 5):
        for col in range(1, len(df.columns) + 2):
            ref = f"{get_column_letter(col)}{row}"
            ws[ref].fill = PatternFill("solid", start_color=DARK, end_color=DARK)

    # Column config: (header, width, format, color_fn)
    COLS = [
        ("Timestamp",    22, None,         None),
        ("Symbol",        8, None,         None),
        ("Spot",         10, '#,##0.00',   None),
        ("Strike",       10, '#,##0.00',   None),
        ("DTE",           8, '0.0',        None),
        ("Bucket",       12, None,         None),
        ("GEX Call",     14, '#,##0',      "GEX_call"),
        ("GEX Put",      14, '#,##0',      "GEX_put"),
        ("GEX Net",      14, '#,##0',      "GEX_net"),
        ("Vanna Call",   14, '#,##0',      "VannEX_call"),
        ("Vanna Put",    14, '#,##0',      "VannEX_put"),
        ("Vanna Net",    14, '#,##0',      "VannEX_net"),
        ("Charm Call",   14, '0.0000',     "CharmEX_call"),
        ("Charm Put",    14, '0.0000',     "CharmEX_put"),
        ("Charm Net",    14, '0.0000',     "CharmEX_net"),
        ("Total OI",     12, '#,##0',      None),
        ("Volume",       12, '#,##0',      None),
        ("IV Call",      10, '0.00%',      None),
        ("IV Put",       10, '0.00%',      None),
    ]

    db_cols = [
        "timestamp","symbol","spot","strike","dte","dte_bucket",
        "GEX_call","GEX_put","GEX_net",
        "VannEX_call","VannEX_put","VannEX_net",
        "CharmEX_call","CharmEX_put","CharmEX_net",
        "total_oi","total_volume","iv_call","iv_put"
    ]

    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 22

    for col_i, (hdr, width, nf, color_col) in enumerate(COLS):
        col_letter = get_column_letter(col_i + 2)
        ws.column_dimensions[col_letter].width = width
        hdr_cell(ws, f"{col_letter}2", hdr, bg=HDR_BG, size=9)

    for row_i, (_, row) in enumerate(df.iterrows()):
        r = row_i + 3
        ws.row_dimensions[r].height = 18
        bg = ALT_BG if row_i % 2 == 0 else DARK

        for col_i, (hdr, width, nf, color_col) in enumerate(COLS):
            col_letter = get_column_letter(col_i + 2)
            val = row[db_cols[col_i]] if db_cols[col_i] in row else None

            fc = WHITE
            if color_col and val is not None:
                fc = color_for_value(val)

            data_cell(ws, f"{col_letter}{r}", val,
                      fc=fc, bg=bg, align="right" if col_i > 1 else "left",
                      nf=nf, size=9)

    ws.freeze_panes = "B3"

def write_summary_sheet(wb, df):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].hidden = True

    for row in range(1, len(df) + 5):
        for col in range(1, len(df.columns) + 2):
            ws[f"{get_column_letter(col)}{row}"].fill = PatternFill("solid", start_color=DARK, end_color=DARK)

    COLS = [
        ("Timestamp",     22, None,          None),
        ("Symbol",         8, None,          None),
        ("Spot",          10, '#,##0.00',    None),
        ("Net GEX",       16, '#,##0',       "net_GEX"),
        ("Net Vanna",     16, '#,##0',       "net_VannEX"),
        ("Net Charm",     16, '0.0000',      "net_CharmEX"),
        ("Gamma Flip",    12, '#,##0.00',    None),
        ("Call Wall",     12, '#,##0.00',    None),
        ("Put Wall",      12, '#,##0.00',    None),
        ("Max Pain",      12, '#,##0.00',    None),
        ("Total OI",      12, '#,##0',       None),
        ("Volume",        12, '#,##0',       None),
        ("ATM IV",        10, '0.00%',       None),
        ("GEX 0DTE",      12, '#,##0',       "gex_0dte"),
        ("GEX 1-7DTE",    12, '#,##0',       "gex_1_7dte"),
        ("GEX 8-30DTE",   12, '#,##0',       "gex_8_30dte"),
        ("GEX 31-90DTE",  12, '#,##0',       "gex_31_90dte"),
        ("GEX 91-180DTE", 12, '#,##0',       "gex_91_180dte"),
        ("GEX 180+DTE",   12, '#,##0',       "gex_180plus_dte"),
    ]

    db_cols = [
        "timestamp","symbol","spot",
        "net_GEX","net_VannEX","net_CharmEX",
        "gamma_flip","call_wall","put_wall","max_pain",
        "total_oi","total_volume","iv_atm",
        "gex_0dte","gex_1_7dte","gex_8_30dte",
        "gex_31_90dte","gex_91_180dte","gex_180plus_dte"
    ]

    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 22

    for col_i, (hdr, width, nf, color_col) in enumerate(COLS):
        col_letter = get_column_letter(col_i + 2)
        ws.column_dimensions[col_letter].width = width
        hdr_cell(ws, f"{col_letter}2", hdr, bg=HDR_BG, size=9)

    for row_i, (_, row) in enumerate(df.iterrows()):
        r = row_i + 3
        ws.row_dimensions[r].height = 18
        bg = ALT_BG if row_i % 2 == 0 else DARK

        for col_i, (hdr, width, nf, color_col) in enumerate(COLS):
            col_letter = get_column_letter(col_i + 2)
            val = row[db_cols[col_i]] if db_cols[col_i] in row else None

            fc = WHITE
            if color_col and val is not None:
                fc = color_for_value(val)

            data_cell(ws, f"{col_letter}{r}", val,
                      fc=fc, bg=bg, align="right" if col_i > 1 else "left",
                      nf=nf, size=9)

    ws.freeze_panes = "B3"

# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print("\n" + "═" * 55)
    print("  Greeks Database Query Tool")
    print("═" * 55)

    validate_config()

    symbols = SYMBOLS if SYMBOLS else []
    conn = sqlite3.connect(DB_PATH)
    print(f"\n  Connected: {DB_PATH}")

    # Resolve timestamps
    timestamps = get_timestamps(conn, symbols)
    if not timestamps:
        print("\n  No data found for your filters. Try adjusting:")
        print("    - DATE_FROM / DATE_TO")
        print("    - SYMBOLS")
        print("    - DTE_BUCKET / DTE_LOW / DTE_HIGH")
        sys.exit(0)

    print(f"  Timestamps matched : {len(timestamps)}")
    for ts in timestamps[:5]:
        print(f"    {ts}")
    if len(timestamps) > 5:
        print(f"    ... and {len(timestamps) - 5} more")

    df_strike  = None
    df_summary = None

    if EXPORT_TABLE in ("strike", "both"):
        df_strike = query_strike_data(conn, timestamps, symbols)
        print(f"\n  Strike rows  : {len(df_strike):,}")

    if EXPORT_TABLE in ("summary", "both"):
        df_summary = query_summary_data(conn, timestamps, symbols)
        print(f"  Summary rows : {len(df_summary):,}")

    conn.close()

    # Build Excel — Strike Data first, Summary second, Query Info third
    wb = Workbook()
    # Remove the default empty sheet openpyxl creates
    wb.remove(wb.active)

    if df_strike is not None and not df_strike.empty:
        write_strike_sheet(wb, df_strike)
    if df_summary is not None and not df_summary.empty:
        write_summary_sheet(wb, df_summary)
    write_title_sheet(wb, symbols, timestamps, df_strike, df_summary)

    out = build_filename()
    wb.save(out)
    print(f"\n  Saved: {out}")
    print("═" * 55 + "\n")

if __name__ == "__main__":
    main()
